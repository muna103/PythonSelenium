import openpyxl
class HomePageData:
    test_HomePage_data = [{"firstName":"muna", "lastName":"R", "gender":"Male"},{"firstName":"sailu ", "lastName":"R", "gender":"Female"}]

    @staticmethod
    def getTestdata(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("F:\Muna\AutomationTesting_Python\PythonDemo.xlsx")
        sheet = book.active
        for i in range(2,sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2,sheet.max_column+1):
                    Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
        return [Dict]